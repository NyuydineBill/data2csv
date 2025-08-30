import csv
import json
import logging
from typing import List, Dict, Any, Optional, Union
from io import StringIO, BytesIO
from django.db import models
from django.conf import settings
from django.utils.translation import gettext_lazy as _
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)


class ExportFormatter:
    """Base formatter for exports."""
    
    def __init__(self, include_headers: bool = True, field_labels: Optional[Dict[str, str]] = None):
        self.include_headers = include_headers
        self.field_labels = field_labels or {}
    
    def format_value(self, value: Any) -> str:
        """Formats a value for export."""
        if value is None:
            return ''
        elif isinstance(value, bool):
            return 'Yes' if value else 'No'
        elif isinstance(value, (models.Model,)):
            return str(value)
        elif isinstance(value, (list, tuple)):
            return ', '.join(str(item) for item in value)
        else:
            return str(value)


class CSVFormatter(ExportFormatter):
    """CSV formatter."""
    
    def __init__(self, delimiter: str = ',', **kwargs):
        super().__init__(**kwargs)
        self.delimiter = delimiter
    
    def export(self, queryset: models.QuerySet, fields: List[str]) -> BytesIO:
        """Exports queryset to CSV."""
        output = StringIO()
        writer = csv.writer(output, delimiter=self.delimiter)
        
        if self.include_headers:
            headers = [self.field_labels.get(field, field) for field in fields]
            writer.writerow(headers)
        
        for obj in queryset:
            row = []
            for field in fields:
                value = self._get_field_value(obj, field)
                row.append(self.format_value(value))
            writer.writerow(row)
        
        output.seek(0)
        return BytesIO(output.getvalue().encode('utf-8'))
    
    def _get_field_value(self, obj: models.Model, field_name: str) -> Any:
        """Gets field value from model."""
        try:
            field = obj._meta.get_field(field_name)
            
            if isinstance(field, models.ForeignKey):
                related_obj = getattr(obj, field_name)
                return related_obj
            elif isinstance(field, models.ManyToManyField):
                return list(getattr(obj, field_name).all())
            else:
                return getattr(obj, field_name)
        except Exception as e:
            logger.warning(f"Error getting field {field_name}: {e}")
            return None


class ExcelFormatter(ExportFormatter):
    """Format exports to Excel."""
    
    def __init__(self, sheet_name: str = 'Sheet1', **kwargs):
        super().__init__(**kwargs)
        self.sheet_name = sheet_name
    
    def export(self, queryset: models.QuerySet, fields: List[str]) -> BytesIO:
        """Export queryset to Excel format."""
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = self.sheet_name
        
        # Style the header row
        header_font = Font(bold=True)
        header_fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        if self.include_headers:
            for col, field in enumerate(fields, 1):
                cell = worksheet.cell(row=1, column=col, value=self.field_labels.get(field, field))
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
        
        # Write data rows
        start_row = 2 if self.include_headers else 1
        for row_idx, obj in enumerate(queryset, start_row):
            for col_idx, field in enumerate(fields, 1):
                value = self._get_field_value(obj, field)
                worksheet.cell(row=row_idx, column=col_idx, value=self.format_value(value))
        
        # Auto-adjust column widths
        self._adjust_column_widths(worksheet, fields)
        
        # Save to BytesIO
        output = BytesIO()
        workbook.save(output)
        output.seek(0)
        return output
    
    def _get_field_value(self, obj: models.Model, field_name: str) -> Any:
        """Get field value from model instance."""
        try:
            field = obj._meta.get_field(field_name)
            
            if isinstance(field, models.ForeignKey):
                related_obj = getattr(obj, field_name)
                return related_obj
            elif isinstance(field, models.ManyToManyField):
                return list(getattr(obj, field_name).all())
            else:
                return getattr(obj, field_name)
        except Exception as e:
            logger.warning(f"Error getting field {field_name}: {e}")
            return None
    
    def _adjust_column_widths(self, worksheet, fields: List[str]):
        """Auto-adjust column widths based on content."""
        for column in worksheet.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            
            for cell in column:
                try:
                    if cell.value:
                        cell_length = len(str(cell.value))
                        max_length = max(max_length, cell_length)
                except:
                    pass
            
            # Set reasonable width (min 10, max 50)
            adjusted_width = max(10, min(max_length + 2, 50))
            worksheet.column_dimensions[column_letter].width = adjusted_width


class JSONFormatter(ExportFormatter):
    """Format exports to JSON."""
    
    def export(self, queryset: models.QuerySet, fields: List[str]) -> BytesIO:
        """Export queryset to JSON format."""
        data = []
        
        for obj in queryset:
            obj_data = {}
            for field in fields:
                value = self._get_field_value(obj, field)
                obj_data[field] = self.format_value(value)
            data.append(obj_data)
        
        json_string = json.dumps(data, indent=2, ensure_ascii=False, default=str)
        return BytesIO(json_string.encode('utf-8'))
    
    def _get_field_value(self, obj: models.Model, field_name: str) -> Any:
        """Get field value from model instance."""
        try:
            field = obj._meta.get_field(field_name)
            
            if isinstance(field, models.ForeignKey):
                related_obj = getattr(obj, field_name)
                return related_obj
            elif isinstance(field, models.ManyToManyField):
                return list(getattr(obj, field_name).all())
            else:
                return getattr(obj, field_name)
        except Exception as e:
            logger.warning(f"Error getting field {field_name}: {e}")
            return None


def get_formatter(format_type: str, **kwargs) -> ExportFormatter:
    """Get the appropriate formatter for the given format type."""
    formatters = {
        'csv': CSVFormatter,
        'xlsx': ExcelFormatter,
        'json': JSONFormatter,
    }
    
    formatter_class = formatters.get(format_type.lower())
    if not formatter_class:
        raise ValueError(f"Unsupported format type: {format_type}")
    
    return formatter_class(**kwargs)


def validate_export_config(config: Dict[str, Any]) -> Dict[str, Any]:
    """Validate export configuration and set defaults."""
    validated = config.copy()
    
    # Set defaults
    defaults = {
        'format': 'csv',
        'include_headers': True,
        'delimiter': getattr(settings, 'EXPORT_CSV_DELIMITER', ','),
        'max_rows': getattr(settings, 'EXPORT_MAX_ROWS', 10000),
        'sheet_name': getattr(settings, 'EXPORT_EXCEL_SHEET_NAME', 'Sheet1'),
    }
    
    for key, default_value in defaults.items():
        if key not in validated:
            validated[key] = default_value
    
    # Validate format-specific options
    if validated['format'] != 'csv':
        validated['delimiter'] = None
    
    return validated


def get_export_filename(base_name: str, format_type: str, timestamp: Optional[str] = None) -> str:
    """Generate export filename with optional timestamp."""
    if timestamp:
        return f"{base_name}_{timestamp}.{format_type}"
    else:
        return f"{base_name}.{format_type}" 