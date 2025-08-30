from django.contrib import admin
from django.http import HttpResponse, JsonResponse
from django.core.exceptions import ValidationError
from django.conf import settings
from django.utils.translation import gettext_lazy as _
from django.contrib import messages
from django.template.response import TemplateResponse
from django.urls import path
from django.utils.safestring import mark_safe
from django.db import models
from typing import List, Dict, Any, Optional, Union
import csv
import json
import logging
from io import StringIO, BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
import mimetypes
import os

logger = logging.getLogger(__name__)

# Configuration defaults
DEFAULT_CONFIG = {
    'CSV_DELIMITER': getattr(settings, 'EXPORT_CSV_DELIMITER', ','),
    'EXCEL_SHEET_NAME': getattr(settings, 'EXPORT_EXCEL_SHEET_NAME', 'Sheet1'),
    'MAX_EXPORT_ROWS': getattr(settings, 'EXPORT_MAX_ROWS', 10000),
    'ENABLE_PROGRESS': getattr(settings, 'EXPORT_ENABLE_PROGRESS', True),
    'DEFAULT_FILENAME': getattr(settings, 'EXPORT_DEFAULT_FILENAME', 'export'),
    'INCLUDE_META_FIELDS': getattr(settings, 'EXPORT_INCLUDE_META_FIELDS', False),
}

class ExportMixin:
    """Mixin that adds export functionality to Django admin models."""
    
    export_fields: Optional[List[str]] = None
    export_exclude_fields: List[str] = []
    export_field_labels: Dict[str, str] = {}
    export_filename: Optional[str] = None
    export_max_rows: Optional[int] = None
    
    enable_csv_export: bool = True
    enable_excel_export: bool = True
    enable_json_export: bool = False
    
    def get_export_fields(self, request) -> List[str]:
        """Returns fields to export."""
        if self.export_fields is not None:
            return self.export_fields
        
        model_fields = [field.name for field in self.model._meta.fields]
        return [field for field in model_fields if field not in self.export_exclude_fields]
    
    def get_export_field_labels(self, request) -> Dict[str, str]:
        """Returns field labels for export."""
        labels = {}
        for field_name in self.get_export_fields(request):
            if field_name in self.export_field_labels:
                labels[field_name] = self.export_field_labels[field_name]
            else:
                field = self.model._meta.get_field(field_name)
                labels[field_name] = getattr(field, 'verbose_name', field_name)
        return labels
    
    def get_export_filename(self, request, format_type: str) -> str:
        """Creates export filename."""
        if self.export_filename:
            base_name = self.export_filename
        else:
            base_name = f"{self.model._meta.verbose_name_plural.lower().replace(' ', '_')}"
        
        timestamp = request.user.date_joined.strftime('%Y%m%d_%H%M%S') if hasattr(request.user, 'date_joined') else ''
        return f"{base_name}_{timestamp}.{format_type}"
    
    def get_export_queryset(self, request, queryset) -> models.QuerySet:
        """Limits queryset to max rows if needed."""
        max_rows = self.export_max_rows or DEFAULT_CONFIG['MAX_EXPORT_ROWS']
        if queryset.count() > max_rows:
            messages.warning(
                request,
                f"Export limited to {max_rows} rows. Use filters to reduce the dataset."
            )
            return queryset[:max_rows]
        return queryset
    
    def export_to_csv(self, request, queryset, *args, **kwargs):
        """Exports to CSV format."""
        try:
            fields = self.get_export_fields(request)
            field_labels = self.get_export_field_labels(request)
            export_queryset = self.get_export_queryset(request, queryset)
            
            response = HttpResponse(content_type='text/csv; charset=utf-8')
            filename = self.get_export_filename(request, 'csv')
            response['Content-Disposition'] = f'attachment; filename="{filename}"'
            
            response.write('\ufeff')
            
            writer = csv.writer(response, delimiter=DEFAULT_CONFIG['CSV_DELIMITER'])
            
            writer.writerow([field_labels.get(field, field) for field in fields])
            
            for obj in export_queryset:
                row = []
                for field in fields:
                    value = self.get_field_value(obj, field)
                    row.append(value)
                writer.writerow(row)
            
            messages.success(request, f"Successfully exported {export_queryset.count()} items to CSV")
            return response
            
        except Exception as e:
            logger.error(f"CSV export failed: {str(e)}")
            messages.error(request, f"Export failed: {str(e)}")
            return HttpResponse(status=500)
    
    def export_to_excel(self, request, queryset, *args, **kwargs):
        """Exports to Excel format."""
        try:
            fields = self.get_export_fields(request)
            field_labels = self.get_export_field_labels(request)
            export_queryset = self.get_export_queryset(request, queryset)
            
            response = HttpResponse(
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            filename = self.get_export_filename(request, 'xlsx')
            response['Content-Disposition'] = f'attachment; filename="{filename}"'
            
            workbook = Workbook()
            worksheet = workbook.active
            worksheet.title = DEFAULT_CONFIG['EXCEL_SHEET_NAME']
            
            header_font = Font(bold=True)
            header_fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
            
            for col, field in enumerate(fields, 1):
                cell = worksheet.cell(row=1, column=col, value=field_labels.get(field, field))
                cell.font = header_font
                cell.fill = header_fill
            
            for row_idx, obj in enumerate(export_queryset, 2):
                for col_idx, field in enumerate(fields, 1):
                    value = self.get_field_value(obj, field)
                    worksheet.cell(row=row_idx, column=col_idx, value=value)
            
            for column in worksheet.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                worksheet.column_dimensions[column_letter].width = adjusted_width
            
            workbook.save(response)
            messages.success(request, f"Successfully exported {export_queryset.count()} items to Excel")
            return response
            
        except Exception as e:
            logger.error(f"Excel export failed: {str(e)}")
            messages.error(request, f"Export failed: {str(e)}")
            return HttpResponse(status=500)
    
    def export_to_json(self, request, queryset, *args, **kwargs):
        """Exports to JSON format."""
        try:
            fields = self.get_export_fields(request)
            export_queryset = self.get_export_queryset(request, queryset)
            
            data = []
            for obj in export_queryset:
                obj_data = {}
                for field in fields:
                    obj_data[field] = self.get_field_value(obj, field)
                data.append(obj_data)
            
            response = HttpResponse(
                json.dumps(data, indent=2, ensure_ascii=False, default=str),
                content_type='application/json; charset=utf-8'
            )
            filename = self.get_export_filename(request, 'json')
            response['Content-Disposition'] = f'attachment; filename="{filename}"'
            
            messages.success(request, f"Successfully exported {export_queryset.count()} items to JSON")
            return response
            
        except Exception as e:
            logger.error(f"JSON export failed: {str(e)}")
            messages.error(request, f"Export failed: {str(e)}")
            return HttpResponse(status=500)
    
    def get_field_value(self, obj, field_name: str) -> Any:
        """Get the value of a field from an object, handling special cases."""
        try:
            field = self.model._meta.get_field(field_name)
            
            # Handle different field types
            if isinstance(field, models.ForeignKey):
                related_obj = getattr(obj, field_name)
                return str(related_obj) if related_obj else ''
            elif isinstance(field, models.ManyToManyField):
                related_objs = getattr(obj, field_name).all()
                return ', '.join(str(related_obj) for related_obj in related_objs)
            elif isinstance(field, models.DateTimeField):
                value = getattr(obj, field_name)
                return value.strftime('%Y-%m-%d %H:%M:%S') if value else ''
            elif isinstance(field, models.DateField):
                value = getattr(obj, field_name)
                return value.strftime('%Y-%m-%d') if value else ''
            elif isinstance(field, models.BooleanField):
                value = getattr(obj, field_name)
                return 'Yes' if value else 'No'
            else:
                return getattr(obj, field_name)
                
        except Exception as e:
            logger.warning(f"Error getting field {field_name} value: {str(e)}")
            return ''
    
    def get_export_actions(self) -> List[str]:
        """Get list of export action names."""
        actions = []
        if self.enable_csv_export:
            actions.append('export_to_csv')
        if self.enable_excel_export:
            actions.append('export_to_excel')
        if self.enable_json_export:
            actions.append('export_to_json')
        return actions
    
    def get_actions(self, request):
        """Override to add export actions."""
        actions = super().get_actions(request)
        
        # Add export actions
        for action_name in self.get_export_actions():
            if action_name == 'export_to_csv':
                actions[action_name] = (self.export_to_csv, action_name, _("Export selected items to CSV"))
            elif action_name == 'export_to_excel':
                actions[action_name] = (self.export_to_excel, action_name, _("Export selected items to Excel"))
            elif action_name == 'export_to_json':
                actions[action_name] = (self.export_to_json, action_name, _("Export selected items to JSON"))
        
        return actions

# Legacy functions for backward compatibility
def export_to_csv(modeladmin, request, queryset):
    """Legacy export function for backward compatibility."""
    if hasattr(modeladmin, 'export_to_csv'):
        return modeladmin.export_to_csv(request, queryset)
    
    # Fallback to basic export
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="exported_data.csv"'
    
    writer = csv.writer(response)
    fields = [field.name for field in queryset.model._meta.fields]
    writer.writerow(fields)
    
    for obj in queryset:
        writer.writerow([getattr(obj, field.name) for field in queryset.model._meta.fields])
    
    return response

def export_to_excel(modeladmin, request, queryset):
    """Legacy export function for backward compatibility."""
    if hasattr(modeladmin, 'export_to_excel'):
        return modeladmin.export_to_excel(request, queryset)
    
    # Fallback to basic export
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="exported_data.xlsx"'
    
    workbook = Workbook()
    worksheet = workbook.active
    
    fields = [field.name for field in queryset.model._meta.fields]
    worksheet.append(fields)
    
    for obj in queryset:
        row = [getattr(obj, field) for field in fields]
        worksheet.append(row)
    
    workbook.save(response)
    return response

# Set descriptions for legacy functions
export_to_csv.short_description = _("Export selected items to CSV")
export_to_excel.short_description = _("Export selected items to Excel")
